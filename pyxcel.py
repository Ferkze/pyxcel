from openpyxl import Workbook
arquivo_excel = Workbook()

planilha1 = arquivo_excel.active
planilha1.title = 'Gastos'

planilha2 = arquivo_excel.create_sheet('Ganhos', 0)

print(arquivo_excel.sheetnames)

planilha1['A1'] = 'Categoria'
planilha1['B1'] = 'Valor'
planilha1['A2'] = 'Restaurante'
planilha1['B2'] = 42.99

valores = [
    ("Categoria", "Valor"),
    ("Restaurante", 45.99),
    ("Transporte", 208.45),
    ("Viagem", 558.54)
]
for linha in valores:
    planilha1.append(linha)

planilha1.cell(row=3, column=1, value=34.99)

planilha1['C1'] = '=SOMA(23,5)'

c1 = planilha1['C1']
print(c1.value)
a1 = planilha1.cell(column=1, row=1)
print(a1.value)

max_linha = planilha1.max_row
max_coluna = planilha1.max_column
for i in range(1, max_linha+1):
    for j in range(1, max_coluna+1):
        print(planilha1.cell(row=i, column=j).value, end=' - ')

arquivo_excel.save('relatorio.xlsx')

#for opening your own existing sheet
path = '/path/to/sheet.xlsx'
arquivo_excel.load_workbook(path)

original = arquivo_excel.get_sheet_by_name('Gastos')
copia = arquivo_excel.copy_worksheet(copia)
arquivo_excel.save('planilha.xlsx')
